"""
clean_sustainability_tracker.py

Scans the merged "Sustainability Reports" Drive folder (ground truth for
what actually exists) and updates the "full list" sheet of the
Sustainability Report Tracker to reflect it accurately.

DESIGN DECISIONS (flagged for Hanzala's review, not silently assumed):
  1. The "full list" sheet has TWO separate FY_2015-FY_2020 column groups
     with genuinely different existing values. Both are updated to the same,
     Drive-verified status, rather than picking one and ignoring the other -
     this IS the "cleaning" the task asked for, resolving an inconsistency
     rather than leaving it.
  2. Drive presence -> "downloaded". Drive absence -> "not found", UNLESS
     the existing cell already says "no reporting", which is preserved
     (that's confirmed human research, not something a file scan can verify
     or safely overwrite).
  3. Output goes to a NEW file first (never overwrites the original in
     place) so this can be reviewed before it replaces anything shared.
"""

import os
import re
from pathlib import Path
from dotenv import load_dotenv
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import openpyxl

load_dotenv()

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
DRIVE_ROOT_FOLDER_ID = os.getenv("DRIVE_ROOT_FOLDER_ID")
OLD_FOLDER_NAME = "Sustainability Reports"
TRACKER_INPUT_PATH = "data/00_reference/Sustainability Report Tracker.xlsx"  # your local copy
TRACKER_OUTPUT_PATH = "data/00_reference/Sustainability Report Tracker_CLEANED.xlsx"


def get_service():
    creds = None
    token_path = Path("token.json")
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("client_secret.json", SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json())
    return build("drive", "v3", credentials=creds)


def find_subfolder(service, parent_id, name):
    query = f"'{parent_id}' in parents and name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    return files[0]["id"] if files else None


def list_ticker_folders(service, parent_id):
    query = f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = []
    page_token = None
    while True:
        response = service.files().list(
            q=query, fields="nextPageToken, files(id, name)", pageToken=page_token
        ).execute()
        results.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return results


def list_pdfs_in_folder(service, folder_id):
    query = f"'{folder_id}' in parents and trashed=false"
    results = []
    page_token = None
    while True:
        response = service.files().list(
            q=query, fields="nextPageToken, files(id, name)", pageToken=page_token
        ).execute()
        results.extend([f["name"] for f in response.get("files", []) if f["name"].lower().endswith(".pdf")])
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return results


def extract_years_from_filenames(filenames):
    """A filename can mention more than one year (e.g. '2018-2019.pdf' or
    '2020-2021.pdf' for a combined report) - credit the company with
    'downloaded' for every year mentioned, not just the first."""
    years = set()
    for name in filenames:
        for match in re.finditer(r"(20\d{2})", name):
            years.add(int(match.group(1)))
    return years


def main():
    print("Authenticating with Google Drive...")
    service = get_service()

    old_folder_id = find_subfolder(service, DRIVE_ROOT_FOLDER_ID, OLD_FOLDER_NAME)
    if not old_folder_id:
        print("FATAL: could not locate the Sustainability Reports folder.")
        return

    print("Scanning all ticker folders (ground truth)...")
    ticker_folders = list_ticker_folders(service, old_folder_id)
    print(f"Found {len(ticker_folders)} ticker folders")

    ticker_years = {}
    for i, folder in enumerate(ticker_folders, 1):
        ticker = folder["name"]
        pdfs = list_pdfs_in_folder(service, folder["id"])
        years = extract_years_from_filenames(pdfs)
        ticker_years[ticker] = years

        if i % 20 == 0:
            print(f"  scanned {i}/{len(ticker_folders)}...")

    print(
        f"Drive scan complete: "
        f"{sum(len(y) for y in ticker_years.values())} total (ticker, year) pairs found"
    )

    print(f"\nLoading tracker: {TRACKER_INPUT_PATH}")
    wb = openpyxl.load_workbook(TRACKER_INPUT_PATH)
    ws = wb["full list"]

    header_row = [cell.value for cell in ws[1]]

    fy_column_indices = [
        i for i, h in enumerate(header_row)
        if h and str(h).startswith("FY_")
    ]

    ticker_col = header_row.index("ticker")
    status_col = header_row.index("status")

    print(f"Found {len(fy_column_indices)} FY_ columns (both groups) to update")

    updated_count = 0
    preserved_no_reporting = 0

    # Determine unique FY columns (first occurrence only) for status calculation
    status_year_columns = []
    seen_headers = set()

    for i, h in enumerate(header_row):
        if h and str(h).startswith("FY_") and h not in seen_headers:
            status_year_columns.append((i, int(str(h).replace("FY_", ""))))
            seen_headers.add(h)

    for row_num in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row_num, column=ticker_col + 1).value

        if not ticker or not isinstance(ticker, str):
            continue

        ticker = ticker.strip()
        drive_years = ticker_years.get(ticker, set())

        # -------------------------
        # Update all FY cells
        # -------------------------
        for col_idx in fy_column_indices:
            header = header_row[col_idx]
            year = int(str(header).replace("FY_", ""))

            cell = ws.cell(row=row_num, column=col_idx + 1)
            current_value = str(cell.value).strip().lower() if cell.value else ""

            if year in drive_years:
                new_value = "downloaded"
            elif current_value == "no reporting":
                new_value = "no reporting"
                preserved_no_reporting += 1
            else:
                new_value = "not found"

            if current_value != new_value:
                cell.value = new_value
                updated_count += 1

        # -------------------------
        # Update STATUS column
        # -------------------------
        downloaded_count = sum(
            year in drive_years
            for _, year in status_year_columns
        )

        total_years = len(status_year_columns)

        if downloaded_count == total_years:
            status = "all found"
        elif downloaded_count == 0:
            status = "not found"
        else:
            status = "partial"

        ws.cell(row=row_num, column=status_col + 1).value = status

    wb.save(TRACKER_OUTPUT_PATH)

    print(f"\n{'=' * 60}")
    print(f"Cells updated: {updated_count}")
    print(f"'no reporting' values preserved: {preserved_no_reporting}")
    print(f"Saved to: {TRACKER_OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()